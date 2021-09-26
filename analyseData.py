from openpyxl import load_workbook

DEST_FILE_NAME = 'employmentsurvey.xlsx'
RESULT_SHEET = "Results"
NO_OF_EMPLOYEES = 40


def get_sheet_data(work_sheet):
    """Returns the sheet data and the headings in the sheet
       by taking worksheet object as input"""
    sheet_data = {}
    headings = [c.value for c in list(work_sheet.rows)[0][1:]]
    for row in list(work_sheet.rows)[1:]:
        sheet_data[row[0].value] = [str(c.value) for c in row[1:]]
    return headings, sheet_data


def categorize_data(headings, engagement_data):
    """Categorize the sheet data based on the headings and
       returns the result data as dict
    """
    i = 0
    result_data = {}
    for heading in headings:
        result = []
        for data in engagement_data.values():
            result.append(data[i])
            result_data[heading] = result
        i = i+1
    return result_data


def analyse_data(result_data):
    """Analyse the data in the result_data dict by
       calculating the average and percentage of the values
       and return the percentage value as analyse_data dict
    """
    analyse_data = {}
    for key, value in result_data.items():
        value = list(map(int, value))
        average = sum(value)/len(value)
        percentage = round(100 * float(average)/float(NO_OF_EMPLOYEES), 2)
        analyse_data[key] = f'{str(percentage)}%'
    return analyse_data


def update_worksheet_result(wb, final_data, headings):
    """create a new worksheet called Result and update the
    final_data values to the sheet
    """
    sheet = wb.create_sheet(RESULT_SHEET)
    headings.insert(0, "Categorisation")
    sheet.append(headings)
    for key, value in final_data.items():
        row_values = list(value.values())
        row_values.insert(0, key)
        sheet.append(row_values)
    wb.save(DEST_FILE_NAME)


def main():
    wb = load_workbook(filename=DEST_FILE_NAME)
    final_data = {}
    for sheet in wb.worksheets:
        if sheet.title == RESULT_SHEET:
            wb.remove(sheet)
            wb.save(DEST_FILE_NAME)
        else:
            headings, sheet_data = get_sheet_data(sheet)
            categorized_data = categorize_data(headings, sheet_data)
            analysed_data = analyse_data(categorized_data)
            final_data[sheet.title] = analysed_data
    update_worksheet_result(wb, final_data, headings)


if __name__ == "__main__":
    main()
