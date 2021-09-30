"""
This module analyses the employment survey details in the excel sheet named
employmentsurvey.xlsx.Excel sheet contains 5 different categories as sheet
and each sheet contains questions the survey results for same. In this
application we read each sheet in the excel sheet and calculate the percentage
of the survey results given that the total no of employees as 40.And update
the survey result of all the sheets to sheet named Results.
Note:As of now we don't get the survey inputs in the application.If you want
to enter the value you can update the excel sheet and run the application.
"""


from openpyxl import load_workbook
import sys
from pprint import pprint

DEST_FILE_NAME = 'employmentsurvey.xlsx'
RESULT_SHEET = "Results"
NO_OF_EMPLOYEES = 40


def get_sheet_data(work_sheet):
    """
    Returns the sheet data and the headings in the sheet
    by taking worksheet object as input
    """
    sheet_data = {}
    print(f"[INFO] Reading survey data for worksheet {work_sheet.title}.")
    # Reading the first row as heading excluding the first value "Questions"
    headings = [c.value for c in list(work_sheet.rows)[0][1:]]
    # Reading the sheet values as sheet data
    for row in list(work_sheet.rows)[1:]:
        sheet_data[row[0].value] = [str(c.value) for c in row[1:]]
    validate_data(sheet_data)
    return headings, sheet_data


def validate_data(sheet_data):
    """
    Inside the try, converts all string values into integers.
    Raises ValueError if there are any strings in the sheet_data.
    """
    try:
        for data in sheet_data.values():
            # This is a placeholder for validation
            _ = [int(value) for value in data]
    except ValueError as e:
        print(f"Invalid data: {e}, please try again.\n")
        sys.exit(1)


def categorize_data(headings, sheet_data, sheet_name):
    """
    Categorize the sheet data based on the headings and
    returns the result data as dict.
    """
    i = 0
    result_data = {}
    print(f"[INFO] Categorizing the data of {sheet_name} according",
          "to the headings.")
    for heading in headings:
        result = []
        # Getting the sheet data values of under corresponding heading say
        # "strongly disagree" into a list named result.
        for data in sheet_data.values():
            result.append(data[i])
            result_data[heading] = result
        i = i+1
    return result_data


def analyse_data(result_data, sheet_name):
    """Analyse the data in the result_data dict by
       calculating the average and percentage of the values
       and return the percentage value as analyse_data dict
    """
    analyse_data = {}
    print(f"[INFO] Analysing the data of {sheet_name}")
    for key, value in result_data.items():
        value = list(map(int, value))
        # Calculating the average of values under each key
        average = sum(value)/len(value)
        # Calculating the percentage for each category and
        # round to 2 decimal points
        percentage = round(100 * float(average)/float(NO_OF_EMPLOYEES), 2)
        analyse_data[key] = f'{str(percentage)}%'
    return analyse_data


def update_worksheet_result(wb, final_data, headings):
    """create a new worksheet called Result and update the
    final_data values to the sheet
    """
    # Creating a worksheet named Results
    sheet = wb.create_sheet(RESULT_SHEET)
    headings.insert(0, "Categorisation")
    # Adding the headings to the first row in the sheet
    sheet.append(headings)
    print(f"[INFO] Updating the result data to a sheet named {RESULT_SHEET}.")
    for key, value in final_data.items():
        # converting the values in dictionary to list
        # so that can inserted into sheet.
        row_values = list(value.values())
        # Inserting the key to the first column of row eg:"Leadership"
        row_values.insert(0, key)
        sheet.append(row_values)
    # Save the file name to update the values
    wb.save(DEST_FILE_NAME)
    print(f"[INFO] Updated {DEST_FILE_NAME}.")


def is_choice_valid(input_question):
    """
    Returns True if user has entered 'y' or returns False.
    if user has entered invalid choice it keeps asking until
    user enters a valid input.
    """
    user_choice = input(input_question).lower().strip()
    while True:
        if user_choice == 'yes':
            return True
        elif user_choice == 'no':
            return False
        else:
            return is_choice_valid("Please Enter yes or no ")


def get_user_name():
    """
    Gets user name to print a welcome message  to the terminal.
    Validates whether the user input is string if not it keeps asking until
    he enters a valid value.
    """
    while True:
        username = input("Hello! Please Enter your name: ")
        try:
            if(username.strip().isdigit()):
                raise ValueError
            else:
                welcome_msg = f"Hi {username}! Welcome to employment survey "\
                            "Analysis application. This application analyses" \
                            " the employment survey details provided in the" \
                            f" excel sheet {DEST_FILE_NAME} and update the "\
                            "results to the Results Sheet in the same file"
                print(welcome_msg)
            break
        except ValueError as _:
            print("Enter a valid value")


def main():
    get_user_name()
    if(is_choice_valid("Would you like to execute? Yes or No :")):
        wb = load_workbook(filename=DEST_FILE_NAME)
        final_data = {}
        validate_data(final_data)
        for sheet in wb.worksheets:
            if sheet.title == RESULT_SHEET:
                wb.remove(sheet)
                wb.save(DEST_FILE_NAME)
            else:
                headings, sheet_data = get_sheet_data(sheet)
                categorized_data = categorize_data(headings,
                                                   sheet_data,
                                                   sheet.title)
                analysed_data = analyse_data(categorized_data, sheet.title)
                final_data[sheet.title] = analysed_data
        print("Here is the final analysis result of Employment survey")
        pprint(final_data)
        update_worksheet_result(wb, final_data, headings)
    else:
        print("Thanks for visiting the application.")
        sys.exit(0)


if __name__ == "__main__":
    main()
